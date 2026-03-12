# ============================================================
#  backend/employees/route.py
#  Employee directory endpoints  (source: employee_data.xlsx)
# ============================================================
"""
Router Layer: employees

หน้าที่:
- GET  /api/employees        — ดึงรายชื่อพนักงานจาก employee_data.xlsx
- POST /api/employees/upload — อัปโหลด .xlsx เพื่อแทนที่ไฟล์ (Admin only)
- GET  /api/employees/export — ดาวน์โหลด employee_data.xlsx ปัจจุบัน
"""

import io
from pathlib import Path

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse

from ..auth.route import get_current_admin

router = APIRouter(prefix="/api/employees", tags=["employees"])

EMPLOYEES_FILE = Path("employee_data.xlsx")
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5 MB

# คอลัมน์ใน xlsx ที่อ่านได้ → key ที่ส่งให้ frontend
_COL_MAP = {
    "รหัสพนักงาน": "code",
    "ชื่อ": "_firstName",
    "นามสกุล": "_lastName",
    "บริษัท": "company",
    "แผนก": "department",
    "ตำแหน่ง": "position",
}
REQUIRED_COLS = {"รหัสพนักงาน", "ชื่อ", "นามสกุล", "บริษัท", "แผนก"}


def _read_employees() -> list[dict]:
    """อ่าน employee_data.xlsx แปลงเป็น list[dict] พร้อมส่งให้ frontend"""
    if not EMPLOYEES_FILE.exists():
        return []
    wb = openpyxl.load_workbook(EMPLOYEES_FILE, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        raw = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
        if not raw.get("รหัสพนักงาน", ""):
            continue  # ข้ามแถวว่าง
        emp = {
            "code": raw.get("รหัสพนักงาน", ""),
            "fullName": (raw.get("ชื่อ", "") + " " + raw.get("นามสกุล", "")).strip(),
            "department": raw.get("แผนก", ""),
            "company": raw.get("บริษัท", ""),
            "position": raw.get("ตำแหน่ง", ""),
        }
        result.append(emp)
    return result


def _validate_xlsx(content: bytes) -> None:
    """ตรวจว่าไฟล์ที่อัปโหลดมีคอลัมน์ครบก่อนบันทึก"""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        wb.close()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"อ่านไฟล์ xlsx ไม่ได้: {exc}")
    if first_row is None:
        raise HTTPException(status_code=400, detail="ไฟล์ว่าง")
    headers = {str(h).strip() for h in first_row if h is not None}
    missing = REQUIRED_COLS - headers
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"ขาดคอลัมน์: {', '.join(sorted(missing))}",
        )


# ---- GET /api/employees — ไม่ต้อง auth (user.html ใช้ด้วย) ----
@router.get("")
def get_employees():
    return _read_employees()


# ---- GET /api/employees/export — ดาวน์โหลดไฟล์ xlsx ปัจจุบัน (Admin only) ----
@router.get("/export")
def export_employees(_=Depends(get_current_admin)):
    if not EMPLOYEES_FILE.exists():
        raise HTTPException(status_code=404, detail="ไม่พบไฟล์ employee_data.xlsx")
    data = EMPLOYEES_FILE.read_bytes()
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employee_data.xlsx"},
    )


# ---- POST /api/employees/upload — รับ .xlsx แทนที่ไฟล์เดิม (Admin only) ----
@router.post("/upload")
async def upload_employees(file: UploadFile = File(...), _=Depends(get_current_admin)):
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="รองรับเฉพาะไฟล์ .xlsx เท่านั้น")

    content = await file.read()

    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="ไฟล์ใหญ่เกิน 5 MB")

    _validate_xlsx(content)

    EMPLOYEES_FILE.write_bytes(content)

    # นับจำนวนพนักงานที่อัปโหลด
    count = len(_read_employees())
    return {"message": f"อัปเดตรายชื่อพนักงาน {count} คนเรียบร้อย", "count": count}
